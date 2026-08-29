# -*- coding: utf-8 -*-

import time
import warnings
from pathlib import Path

import pandas as pd
import win32com.client as win32
from openpyxl import load_workbook


warnings.simplefilter("ignore", UserWarning)


# ============================================================
# ПУТИ К ФАЙЛАМ
# ============================================================

SUMMARY_PATH = Path(
    r"C:\Users\h301145\OneDrive - Halliburton"
    r"\Desktop\Summary_Reports.xlsm"
)

MTT_PATH = Path(
    r"C:\Users\h301145\OneDrive - Halliburton"
    r"\Rigs Inventory\Execution\MTT Tracker.xlsm"
)

RIGS_PATH = Path(
    r"C:\Users\h301145\OneDrive - Halliburton"
    r"\Rigs Inventory\Execution"
    r"\Rigs Inventory Daily Activity Log.xlsm"
)

ZNAME_PATH = Path(
    r"C:\Users\h301145\OneDrive - Halliburton"
    r"\Desktop\Python\3_zName\zName.xlsx"
)

ZMBL_PATH = Path(
    r"C:\Users\h301145\OneDrive - Halliburton"
    r"\Desktop\ZMBL_1111.xlsx"
)


# ============================================================
# НАЗВАНИЯ ЛИСТОВ
# ============================================================

MTT_WELLS_SHEET = "Wells & SOs"
MTT_DATA_SHEET = "MTT Data"
STORE_DATA_SHEET = "Store Data"
ZNAME_SHEET = "cCDB"
ZMBL_SHEET = "ZMBL"


# ============================================================
# ЛИСТЫ SUMMARY_REPORTS ДЛЯ ОБРАБОТКИ
# ============================================================

SUMMARY_SHEETS = [
    "AD-71",
    "HP-701",
    "SP-258",
    "SP-32",
    "T-901",
    "T-902",

    "HP-704",
    "HP-705",
    "SP-259",
    "SP-260",
    "SP-262",
    "SP-27",

    "AD-73",
    "SP-257",
    "SP-26",
    "SP-261",
    "SP-29",
    "SP-31",

    "AD-74",
    "AD-75",
    "HP-700",
    "SP-30",
    "SP-93",
    "SP-99",

    "AD-72",
    "AD-76",
    "AD-77",
    "AD-78",
    "HP-703",
    "HP-707",

    "HP-702",
    "HP-706",
    "NBR-16S",
    "NBR-20S",
]


# ============================================================
# РАСПОЛОЖЕНИЕ КОЛОНОК ДЛЯ ШЕСТИ СКВАЖИН
# ============================================================
#
# I скважина:
#   C = фактическое количество
#   D = MTT
#
# II скважина:
#   E = фактическое количество
#   F = MTT
#
# III скважина:
#   G = фактическое количество
#   H = MTT
#
# IV скважина:
#   I = фактическое количество
#   J = MTT
#
# V скважина:
#   K = фактическое количество
#   L = MTT
#
# VI скважина:
#   M = фактическое количество
#   N = MTT
#
# Дополнительные колонки:
#   O = Used / YTD Consumption
#   Q = Ending / DMR
#   S = ZMBL / LMP
# ============================================================

WELL_COLUMN_PAIRS = {
    "C": "D",
    "E": "F",
    "G": "H",
    "I": "J",
    "K": "L",
    "M": "N",
}

ACTUAL_COLUMNS = list(WELL_COLUMN_PAIRS.keys())
MTT_COLUMNS = list(WELL_COLUMN_PAIRS.values())

ZNAME_USED_COLUMN = "O"
ZNAME_ENDING_COLUMN = "Q"
ZMBL_COLUMN = "S"


# ============================================================
# ПРОВЕРКА НАЛИЧИЯ ФАЙЛОВ
# ============================================================

def check_input_files():
    files = {
        "Summary Reports": SUMMARY_PATH,
        "MTT Tracker": MTT_PATH,
        "Rigs Inventory": RIGS_PATH,
        "zName": ZNAME_PATH,
        "ZMBL": ZMBL_PATH,
    }

    missing_files = []

    for file_name, file_path in files.items():
        if not file_path.exists():
            missing_files.append(
                f"{file_name}: {file_path}"
            )

    if missing_files:
        raise FileNotFoundError(
            "Не найдены следующие файлы:\n"
            + "\n".join(missing_files)
        )


# ============================================================
# АВТОМАТИЧЕСКИЙ ПЕРЕСЧЁТ EXCEL
# ============================================================

def force_excel_recalculation(file_paths):
    """
    Открывает файлы через настоящий Excel,
    полностью пересчитывает формулы и сохраняет результаты.

    Это необходимо, потому что openpyxl самостоятельно
    формулы не рассчитывает.
    """

    excel = None
    opened_workbooks = []

    try:
        print("=" * 60)
        print("Запускаю автоматический пересчёт Excel...")
        print("=" * 60)

        excel = win32.DispatchEx("Excel.Application")

        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        for file_path in file_paths:
            print(
                f"Открываю для пересчёта: "
                f"{file_path.name}"
            )

            workbook = excel.Workbooks.Open(
                Filename=str(file_path),
                UpdateLinks=0,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True
            )

            opened_workbooks.append(workbook)

        # xlCalculationAutomatic
        excel.Calculation = -4105

        print("Выполняю полный пересчёт формул...")

        # Эквивалент Ctrl + Alt + F9.
        excel.CalculateFull()

        timeout_seconds = 300
        start_time = time.time()

        # CalculationState = 0 означает xlDone.
        while excel.CalculationState != 0:
            if time.time() - start_time > timeout_seconds:
                raise TimeoutError(
                    "Excel не завершил пересчёт "
                    "за 5 минут."
                )

            time.sleep(0.5)

        print("Сохраняю пересчитанные файлы...")

        for workbook in opened_workbooks:
            workbook.Save()

        print("Автоматический пересчёт завершён.")

    finally:
        for workbook in reversed(opened_workbooks):
            try:
                workbook.Close(SaveChanges=True)
            except Exception:
                pass

        if excel is not None:
            try:
                excel.EnableEvents = True
                excel.ScreenUpdating = True
                excel.DisplayAlerts = True
                excel.Quit()
            except Exception:
                pass


# ============================================================
# ОПРЕДЕЛЕНИЕ СТРОК ЗАГОЛОВКОВ
# ============================================================

def detect_header_row_mtt(path, sheet):
    temp = pd.read_excel(
        path,
        sheet_name=sheet,
        header=None,
        nrows=30
    )

    for i in range(len(temp)):
        row = (
            temp.iloc[i]
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )

        if (
            "well" in row
            and "product" in row
            and (
                "mtt qty" in row
                or ("mtt" in row and "qty" in row)
            )
        ):
            return i

    raise Exception(
        "Не удалось определить строку заголовков "
        "на листе 'MTT Data'."
    )


def detect_header_row_store(path, sheet):
    temp = pd.read_excel(
        path,
        sheet_name=sheet,
        header=None,
        nrows=30
    )

    for i in range(len(temp)):
        row = (
            temp.iloc[i]
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )

        if (
            "sloc" in row
            and "material description" in row
            and "qty" in row
        ):
            return i

    raise Exception(
        "Не удалось определить строку заголовков "
        "на листе 'Store Data'."
    )


def detect_header_row_zname(path, sheet):
    temp = pd.read_excel(
        path,
        sheet_name=sheet,
        header=None,
        nrows=30
    )

    for i in range(len(temp)):
        row = (
            temp.iloc[i]
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )

        if (
            "well name" in row
            and "product name" in row
        ):
            return i

    raise Exception(
        "Не удалось определить строку заголовков "
        "в zName.xlsx."
    )


def detect_header_row_zmbl(path, sheet):
    temp = pd.read_excel(
        path,
        sheet_name=sheet,
        header=None,
        nrows=30
    )

    for i in range(len(temp)):
        row = (
            temp.iloc[i]
            .astype(str)
            .str.strip()
            .str.lower()
            .tolist()
        )

        if (
            "product name" in row
            and (
                "sloc qty" in row
                or "sloc quantity" in row
            )
        ):
            return i

    raise Exception(
        "Не удалось определить строку заголовков "
        "в ZMBL_1111.xlsx."
    )


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def to_float(value):
    try:
        return float(
            str(value).replace(",", "")
        )
    except (ValueError, TypeError):
        return 0.0


def normalize_text(value):
    if value is None:
        return ""

    result = str(value).strip()

    if result.lower() in (
        "",
        "nan",
        "none",
        "nat",
    ):
        return ""

    return result


def normalize_stock_name(value):
    return normalize_text(value).upper()


def is_empty_excel_value(value):
    result = normalize_text(value).lower()

    return result in (
        "",
        "nan",
        "none",
        "nat",
    )


def is_z_material(material_name):
    material_name = normalize_text(material_name)

    return (
        len(material_name) > 1
        and material_name.lower().startswith("z")
    )


def strip_leading_z(material_name):
    material_name = normalize_text(material_name)

    if is_z_material(material_name):
        return material_name[1:].strip()

    return material_name


def get_displayed_cell_value(
    ws,
    ws_vals,
    cell_address
):
    """
    Возвращает отображаемое значение ячейки.

    Если ячейка содержит формулу, значение берётся
    из книги, открытой с data_only=True.

    Если ячейка содержит обычный текст или число,
    значение берётся из обычной книги.
    """

    raw_value = ws[cell_address].value

    if (
        isinstance(raw_value, str)
        and raw_value.startswith("=")
    ):
        return ws_vals[cell_address].value

    return raw_value


def get_material_value(ws, ws_vals, row):
    """
    Получает отображаемое название материала из колонки A.

    Работает как с обычными текстовыми значениями,
    так и с названиями, полученными формулами.
    """

    return get_displayed_cell_value(
        ws,
        ws_vals,
        f"A{row}"
    )


# ============================================================
# ПОЛУЧЕНИЕ ТЕКУЩЕЙ СКВАЖИНЫ ИЗ A2
# ============================================================

def get_current_well(ws_vals, ws, wb_vals):
    current_value = get_displayed_cell_value(
        ws,
        ws_vals,
        "A2"
    )

    if normalize_text(current_value):
        return normalize_text(current_value)

    # Резервный способ для простой формулы
    # вида =DailySummRep!D32.
    raw_formula = ws["A2"].value

    if (
        isinstance(raw_formula, str)
        and raw_formula.startswith("=")
        and "!" in raw_formula
    ):
        expression = raw_formula[1:]

        sheet_name_ref, cell_ref = expression.split(
            "!",
            1
        )

        sheet_name_ref = (
            sheet_name_ref
            .strip()
            .strip("'")
        )

        cell_ref = (
            cell_ref
            .replace("$", "")
            .strip()
        )

        if sheet_name_ref in wb_vals.sheetnames:
            ws_ref = wb_vals[sheet_name_ref]
            referenced_value = ws_ref[cell_ref].value

            if normalize_text(referenced_value):
                return normalize_text(
                    referenced_value
                )

    return None


# ============================================================
# Wells & SOs
# ============================================================

def build_well_map():
    df = pd.read_excel(
        MTT_PATH,
        sheet_name=MTT_WELLS_SHEET,
        header=3,
        dtype=str
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        "Rig",
        "Well List",
        "Parent SO #",
        "Status",
    ]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(
                f"В листе 'Wells & SOs' "
                f"не найдена колонка '{column}'."
            )

    df = (
        df[required_columns]
        .dropna(how="all")
        .copy()
    )

    well_map = {}

    for _, row in df.iterrows():
        rig = normalize_text(row["Rig"])
        well = normalize_text(row["Well List"])

        parent_so = normalize_text(
            row["Parent SO #"]
        )

        status = normalize_text(
            row["Status"]
        )

        if rig and well:
            key = (rig, well)

            if key not in well_map:
                well_map[key] = (
                    parent_so,
                    status
                )

    print(
        f"Wells & SOs: найдено записей = "
        f"{len(well_map)}"
    )

    return well_map


# ============================================================
# MTT Data
# ============================================================

def load_mtt_data():
    header_row = detect_header_row_mtt(
        MTT_PATH,
        MTT_DATA_SHEET
    )

    df = pd.read_excel(
        MTT_PATH,
        sheet_name=MTT_DATA_SHEET,
        header=header_row,
        dtype=str
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        "Well",
        "Product",
        "MTT QTY",
    ]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(
                f"В листе 'MTT Data' "
                f"не найдена колонка '{column}'."
            )

    rig_column_name = None

    for candidate in [
        "Rig",
        "Rig Name",
    ]:
        if candidate in df.columns:
            rig_column_name = candidate
            break

    use_columns = required_columns.copy()

    if rig_column_name:
        use_columns.append(rig_column_name)

    df = (
        df[use_columns]
        .dropna(how="all")
        .copy()
    )

    df["Well"] = (
        df["Well"]
        .apply(normalize_text)
    )

    df["Product"] = (
        df["Product"]
        .apply(normalize_text)
    )

    df["MTT QTY"] = (
        df["MTT QTY"]
        .apply(to_float)
    )

    if rig_column_name:
        df[rig_column_name] = (
            df[rig_column_name]
            .apply(normalize_text)
        )

    print(
        f"MTT Data: загружено строк = {len(df)}"
    )

    return df, rig_column_name


def build_mtt_qty_map(
    well_name,
    rig_name,
    mtt_df,
    rig_column_name
):
    mask = mtt_df["Well"] == well_name

    if rig_column_name:
        mask = (
            mask
            & (mtt_df[rig_column_name] == rig_name)
        )

    filtered_df = mtt_df[mask].copy()

    grouped = (
        filtered_df
        .groupby("Product")["MTT QTY"]
        .sum()
        .to_dict()
    )

    return {
        normalize_text(product): qty
        for product, qty in grouped.items()
    }


# ============================================================
# Store Data
# ============================================================

def build_store_qty_maps():
    header_row = detect_header_row_store(
        RIGS_PATH,
        STORE_DATA_SHEET
    )

    df = pd.read_excel(
        RIGS_PATH,
        sheet_name=STORE_DATA_SHEET,
        header=header_row,
        dtype=str
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        "Sloc",
        "Material Description",
        "QTY",
        "LMP VR Number",
        "Stock Name",
        "Fluids System",
    ]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(
                f"В листе 'Store Data' "
                f"не найдена колонка '{column}'."
            )

    df = (
        df[required_columns]
        .dropna(how="all")
        .copy()
    )

    df["Sloc"] = (
        df["Sloc"]
        .apply(normalize_text)
    )

    df["Material Description"] = (
        df["Material Description"]
        .apply(normalize_text)
    )

    df["QTY"] = (
        df["QTY"]
        .apply(to_float)
    )

    df["LMP VR Number"] = (
        df["LMP VR Number"]
        .apply(normalize_text)
    )

    df["Stock Name"] = (
        df["Stock Name"]
        .apply(normalize_stock_name)
    )

    df["Fluids System"] = (
        df["Fluids System"]
        .apply(
            lambda value:
            normalize_text(value).upper()
        )
    )

    # Оставляем строки без LMP VR Number.
    df = df[
        df["LMP VR Number"].apply(
            is_empty_excel_value
        )
    ].copy()

    # Исключаем Cement Job и Aramco Spacer.
    df = df[
        ~df["Fluids System"].isin(
            [
                "CEMENT JOB",
                "ARAMCO SPACER",
            ]
        )
    ].copy()

    regular_df = df[
        df["Stock Name"] != "ARAMCO"
    ].copy()

    aramco_df = df[
        df["Stock Name"] == "ARAMCO"
    ].copy()

    regular_grouped = (
        regular_df
        .groupby(
            [
                "Sloc",
                "Material Description",
            ]
        )["QTY"]
        .sum()
    )

    aramco_grouped = (
        aramco_df
        .groupby(
            [
                "Sloc",
                "Material Description",
            ]
        )["QTY"]
        .sum()
    )

    regular_store_map = {}

    for (sloc, material), qty in regular_grouped.items():
        key = (
            normalize_text(sloc),
            normalize_text(material),
        )

        regular_store_map[key] = qty

    aramco_store_map = {}

    for (sloc, material), qty in aramco_grouped.items():
        key = (
            normalize_text(sloc),
            normalize_text(material),
        )

        aramco_store_map[key] = qty

    print(
        "Store Data: "
        f"regular rows = {len(regular_store_map)}"
    )

    print(
        "Store Data: "
        f"ARAMCO rows = {len(aramco_store_map)}"
    )

    return (
        regular_store_map,
        aramco_store_map
    )


# ============================================================
# zName
# ============================================================

def load_zname_data():
    header_row = detect_header_row_zname(
        ZNAME_PATH,
        ZNAME_SHEET
    )

    df = pd.read_excel(
        ZNAME_PATH,
        sheet_name=ZNAME_SHEET,
        header=header_row,
        dtype=str
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        "Well Name",
        "Product Name",
        "Used",
        "Ending",
    ]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(
                f"В zName.xlsx "
                f"не найдена колонка '{column}'."
            )

    df = (
        df[required_columns]
        .dropna(how="all")
        .copy()
    )

    df["Well Name"] = (
        df["Well Name"]
        .apply(normalize_text)
    )

    df["Product Name"] = (
        df["Product Name"]
        .apply(normalize_text)
    )

    df["Used"] = (
        df["Used"]
        .apply(to_float)
    )

    df["Ending"] = (
        df["Ending"]
        .apply(to_float)
    )

    print(
        f"zName: загружено строк = {len(df)}"
    )

    return df


def build_zname_map(
    well_name,
    value_column,
    zname_df
):
    filtered_df = zname_df[
        zname_df["Well Name"] == well_name
    ].copy()

    grouped = (
        filtered_df
        .groupby("Product Name")[value_column]
        .sum()
        .to_dict()
    )

    return {
        normalize_text(product): qty
        for product, qty in grouped.items()
    }


# ============================================================
# ZMBL
# ============================================================

def build_zmbl_map():
    header_row = detect_header_row_zmbl(
        ZMBL_PATH,
        ZMBL_SHEET
    )

    df = pd.read_excel(
        ZMBL_PATH,
        sheet_name=ZMBL_SHEET,
        header=header_row,
        dtype=str
    )

    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = [
        "Product name",
        "Sloc QTY",
    ]

    for column in required_columns:
        if column not in df.columns:
            raise ValueError(
                f"В ZMBL_1111.xlsx "
                f"не найдена колонка '{column}'."
            )

    df = (
        df[required_columns]
        .dropna(how="all")
        .copy()
    )

    df["Product name"] = (
        df["Product name"]
        .apply(normalize_text)
    )

    df["Sloc QTY"] = (
        df["Sloc QTY"]
        .apply(to_float)
    )

    grouped = (
        df
        .groupby("Product name")["Sloc QTY"]
        .sum()
        .to_dict()
    )

    zmbl_map = {
        normalize_text(product): qty
        for product, qty in grouped.items()
    }

    print(
        f"ZMBL: rows = {len(zmbl_map)}"
    )

    return zmbl_map


# ============================================================
# ОЧИСТКА КОЛОНОК ДО FINISH
# ============================================================

def clear_columns_until_finish(
    ws,
    ws_vals,
    columns
):
    row = 6

    while row <= ws.max_row:
        # Получаем отображаемый результат формулы
        # из колонки A.
        material = get_material_value(
            ws,
            ws_vals,
            row
        )

        material_name = normalize_text(material)

        if material_name.upper() == "FINISH":
            break

        if material_name:
            for column in columns:
                ws[f"{column}{row}"].value = None

        row += 1


def clear_so_status_block(ws):
    for column in ACTUAL_COLUMNS:
        ws[f"{column}2"].value = None
        ws[f"{column}3"].value = None


# ============================================================
# ОБНОВЛЕНИЕ SO И STATUS
# ============================================================

def update_so_status_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    well_map
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    rig_name = sheet_name

    clear_so_status_block(ws)

    for column in ACTUAL_COLUMNS:
        well_value = get_displayed_cell_value(
            ws,
            ws_vals,
            f"{column}1"
        )

        well_name = normalize_text(well_value)

        if not well_name:
            continue

        entry = well_map.get(
            (rig_name, well_name)
        )

        # Если Rig + Well не найдены,
        # пробуем найти только по Well.
        if entry is None:
            candidates = [
                value
                for (rig, well), value
                in well_map.items()
                if well == well_name
            ]

            if len(candidates) == 1:
                entry = candidates[0]

        if entry is not None:
            parent_so, status = entry

            ws[f"{column}2"].value = parent_so
            ws[f"{column}3"].value = status

            print(
                f"{sheet_name}: {well_name} → "
                f"SO {parent_so}, Status {status}"
            )
        else:
            print(
                f"{sheet_name}: не найден SO/Status "
                f"для скважины {well_name}."
            )


# ============================================================
# ОБНОВЛЕНИЕ ФАКТА ИЗ STORE DATA
# ============================================================

def update_actual_consumption_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    regular_store_map,
    aramco_store_map
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    clear_columns_until_finish(
        ws,
        ws_vals,
        ACTUAL_COLUMNS
    )

    for column in ACTUAL_COLUMNS:
        sloc_value = ws[f"{column}2"].value
        sloc_code = normalize_text(sloc_value)

        if not sloc_code:
            continue

        row = 6

        while row <= ws.max_row:
            material = get_material_value(
                ws,
                ws_vals,
                row
            )

            material_name = normalize_text(material)

            if material_name.upper() == "FINISH":
                break

            if not material_name:
                row += 1
                continue

            if is_z_material(material_name):
                lookup_name = strip_leading_z(
                    material_name
                )

                key = (
                    sloc_code,
                    lookup_name,
                )

                if key in aramco_store_map:
                    ws[f"{column}{row}"].value = (
                        aramco_store_map[key]
                    )

            else:
                key = (
                    sloc_code,
                    material_name,
                )

                if key in regular_store_map:
                    ws[f"{column}{row}"].value = (
                        regular_store_map[key]
                    )

            row += 1


# ============================================================
# ОБНОВЛЕНИЕ MTT DATA
# ============================================================

def update_mtt_data_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    mtt_df,
    rig_column_name
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    rig_name = sheet_name

    clear_columns_until_finish(
        ws,
        ws_vals,
        MTT_COLUMNS
    )

    for well_column, qty_column in WELL_COLUMN_PAIRS.items():
        well_value = get_displayed_cell_value(
            ws,
            ws_vals,
            f"{well_column}1"
        )

        well_name = normalize_text(well_value)

        if not well_name:
            continue

        mtt_map = build_mtt_qty_map(
            well_name=well_name,
            rig_name=rig_name,
            mtt_df=mtt_df,
            rig_column_name=rig_column_name
        )

        row = 6

        while row <= ws.max_row:
            material = get_material_value(
                ws,
                ws_vals,
                row
            )

            material_name = normalize_text(material)

            if material_name.upper() == "FINISH":
                break

            if not material_name:
                row += 1
                continue

            if material_name in mtt_map:
                ws[f"{qty_column}{row}"].value = (
                    mtt_map[material_name]
                )

            row += 1


# ============================================================
# ОБНОВЛЕНИЕ USED ИЗ zName В КОЛОНКУ O
# ============================================================

def update_zname_used_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    zname_df
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    current_well = get_current_well(
        ws_vals,
        ws,
        wb_vals
    )

    clear_columns_until_finish(
        ws,
        ws_vals,
        [ZNAME_USED_COLUMN]
    )

    if current_well is None:
        print(
            f"{sheet_name}: A2 пустая, "
            f"колонка {ZNAME_USED_COLUMN} пропущена."
        )
        return

    zname_used_map = build_zname_map(
        well_name=current_well,
        value_column="Used",
        zname_df=zname_df
    )

    row = 6

    while row <= ws.max_row:
        material = get_material_value(
            ws,
            ws_vals,
            row
        )

        material_name = normalize_text(material)

        if material_name.upper() == "FINISH":
            break

        if not material_name:
            row += 1
            continue

        if material_name in zname_used_map:
            ws[f"{ZNAME_USED_COLUMN}{row}"].value = (
                zname_used_map[material_name]
            )

        row += 1


# ============================================================
# ОБНОВЛЕНИЕ ENDING ИЗ zName В КОЛОНКУ Q
# ============================================================

def update_zname_ending_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    zname_df
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    current_well = get_current_well(
        ws_vals,
        ws,
        wb_vals
    )

    clear_columns_until_finish(
        ws,
        ws_vals,
        [ZNAME_ENDING_COLUMN]
    )

    if current_well is None:
        print(
            f"{sheet_name}: A2 пустая, "
            f"колонка {ZNAME_ENDING_COLUMN} пропущена."
        )
        return

    zname_ending_map = build_zname_map(
        well_name=current_well,
        value_column="Ending",
        zname_df=zname_df
    )

    row = 6

    while row <= ws.max_row:
        material = get_material_value(
            ws,
            ws_vals,
            row
        )

        material_name = normalize_text(material)

        if material_name.upper() == "FINISH":
            break

        if not material_name:
            row += 1
            continue

        if material_name in zname_ending_map:
            ws[f"{ZNAME_ENDING_COLUMN}{row}"].value = (
                zname_ending_map[material_name]
            )

        row += 1


# ============================================================
# ОБНОВЛЕНИЕ ZMBL В КОЛОНКУ S
# ============================================================

def update_zmbl_to_summary_sheet(
    sheet_name,
    wb,
    wb_vals,
    zmbl_map
):
    if sheet_name not in wb.sheetnames:
        print(
            f"Лист {sheet_name} не найден. "
            "Пропускаю."
        )
        return

    if sheet_name not in wb_vals.sheetnames:
        print(
            f"Лист {sheet_name} не найден "
            "в data_only. Пропускаю."
        )
        return

    ws = wb[sheet_name]
    ws_vals = wb_vals[sheet_name]

    clear_columns_until_finish(
        ws,
        ws_vals,
        [ZMBL_COLUMN]
    )

    row = 6

    while row <= ws.max_row:
        material = get_material_value(
            ws,
            ws_vals,
            row
        )

        material_name = normalize_text(material)

        if material_name.upper() == "FINISH":
            break

        if not material_name:
            row += 1
            continue

        if material_name in zmbl_map:
            ws[f"{ZMBL_COLUMN}{row}"].value = (
                zmbl_map[material_name]
            )

        row += 1


# ============================================================
# ОСНОВНАЯ ФУНКЦИЯ
# ============================================================

def main():
    print("=" * 60)
    print("ОБНОВЛЕНИЕ SUMMARY REPORTS")
    print(
        f"Будет обработано листов: "
        f"{len(SUMMARY_SHEETS)}"
    )
    print("=" * 60)

    check_input_files()

    # Сначала Excel пересчитывает формулы в Summary и MTT.
    # После этого openpyxl получает отображаемые результаты
    # формул через data_only=True.
    force_excel_recalculation(
        [
            SUMMARY_PATH,
            MTT_PATH,
        ]
    )

    print("Открываю Summary_Reports.xlsm...")

    # Книга с формулами.
    wb = load_workbook(
        SUMMARY_PATH,
        data_only=False,
        keep_vba=True
    )

    # Книга с сохранёнными результатами формул.
    wb_vals = load_workbook(
        SUMMARY_PATH,
        data_only=True,
        keep_vba=True
    )

    update_completed = False

    try:
        print("Загружаю Wells & SOs...")
        well_map = build_well_map()

        print("Загружаю MTT Data...")
        mtt_df, rig_column_name = load_mtt_data()

        print("Загружаю Store Data...")

        (
            regular_store_map,
            aramco_store_map
        ) = build_store_qty_maps()

        print("Загружаю zName...")
        zname_df = load_zname_data()

        print("Загружаю ZMBL...")
        zmbl_map = build_zmbl_map()

        processed_count = 0

        for sheet_number, sheet_name in enumerate(
            SUMMARY_SHEETS,
            start=1
        ):
            print("-" * 60)

            print(
                f"[{sheet_number}/{len(SUMMARY_SHEETS)}] "
                f"Начинаю обработку листа {sheet_name}..."
            )

            if sheet_name not in wb.sheetnames:
                print(
                    f"Лист {sheet_name} не найден. "
                    "Полностью пропускаю."
                )
                continue

            update_so_status_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                well_map=well_map
            )

            update_actual_consumption_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                regular_store_map=regular_store_map,
                aramco_store_map=aramco_store_map
            )

            update_mtt_data_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                mtt_df=mtt_df,
                rig_column_name=rig_column_name
            )

            update_zname_used_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                zname_df=zname_df
            )

            update_zname_ending_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                zname_df=zname_df
            )

            update_zmbl_to_summary_sheet(
                sheet_name=sheet_name,
                wb=wb,
                wb_vals=wb_vals,
                zmbl_map=zmbl_map
            )

            processed_count += 1

            print(
                f"Готово: лист {sheet_name} обработан."
            )

        print("-" * 60)
        print("Сохраняю Summary_Reports.xlsm...")

        wb.save(SUMMARY_PATH)

        update_completed = True

        print(
            "Файл Summary_Reports.xlsm "
            "обновлён и сохранён."
        )

        print(
            f"Фактически обработано листов: "
            f"{processed_count}"
        )

    finally:
        wb.close()
        wb_vals.close()

    # После сохранения через openpyxl повторно открываем
    # Summary через настоящий Excel и пересчитываем формулы.
    if update_completed:
        print()
        print(
            "Выполняю финальный пересчёт "
            "Summary_Reports.xlsm..."
        )

        force_excel_recalculation(
            [
                SUMMARY_PATH,
            ]
        )

    print("=" * 60)
    print("ОБНОВЛЕНИЕ ВСЕХ ЛИСТОВ ЗАВЕРШЕНО")
    print("=" * 60)


# ============================================================
# ЗАПУСК
# ============================================================

if __name__ == "__main__":
    try:
        main()

    except PermissionError:
        print()
        print(
            "ОШИБКА: не удалось открыть "
            "или сохранить один из файлов."
        )
        print(
            "Закрой Summary_Reports.xlsm и "
            "MTT Tracker.xlsm в Excel, "
            "затем запусти код повторно."
        )

    except TimeoutError as error:
        print()
        print(
            f"ОШИБКА ПЕРЕСЧЁТА: {error}"
        )

    except Exception as error:
        print()
        print(
            f"ОШИБКА: "
            f"{type(error).__name__}: {error}"
        )

    input("\nНажми Enter для закрытия...")